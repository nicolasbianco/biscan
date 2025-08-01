from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask import send_from_directory
import os
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
from datetime import datetime
import base64
from io import BytesIO
import zipfile

UPLOAD_FOLDER = 'uploads'
EXCEL_FILE = 'activos.xlsx'

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

app.secret_key = "supersegreta"
ADMIN_USERNAME = "nbianco"
ADMIN_PASSWORD = "bitroncina"

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Inicializar el archivo Excel si no existe
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Codigo', 'Fecha', 'RutaFoto'])
        wb.save(EXCEL_FILE)

# Verificar si el código ya existe en el Excel
def codigo_existe(codigo):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == codigo:
            return True
    return False

# Agregar registro al Excel
def guardar_activo(codigo, ruta_foto):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append([codigo, fecha, ruta_foto])
    wb.save(EXCEL_FILE)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/scan", methods=["POST"])
def scan():
    codigo = request.form.get("qr_content", "")
    if not codigo:
        return render_template("error.html", message="No se recibió ningún código QR.")
    init_excel()
    if codigo_existe(codigo):
        return render_template("error.html", message=f"El código {codigo} ya existe en la base de datos.")
    # Si no existe, permitir tomar foto y pasar el código a la siguiente página
    return render_template("foto.html", codigo=codigo)

@app.route("/upload_foto", methods=["POST"])
def upload_foto():
    codigo = request.form.get("codigo", "")
    img_data = request.form.get("foto", "")
    if not codigo or not img_data:
        return render_template("error.html", message="Datos incompletos para guardar el activo.")
    # Decodificar la imagen base64
    header, encoded = img_data.split(",", 1)
    data = base64.b64decode(encoded)
    filename = secure_filename(f"{codigo}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    with open(filepath, "wb") as f:
        f.write(data)
    # Guardar en excel
    guardar_activo(codigo, filepath)
    return render_template("success.html", codigo=codigo, ruta_foto=filepath)

@app.route("/uploads/<filename>")
def uploaded_file(filename):
   return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("admin"))
        flash("Credenziali non valide")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect(url_for("login"))

@app.route("/admin")
def admin():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    files = os.listdir(app.config["UPLOAD_FOLDER"])
    files = [f for f in files if f.lower().endswith((".jpg", ".jpeg", ".png"))]
    return render_template("admin.html", files=files)

if __name__ == "__main__":
    init_excel()
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)
    
